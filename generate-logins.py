#!/usr/bin/env python3
"""
Edubeam LMS -- Bulk Login Generator (Python)
Generates PRINCIPAL, STUDENT and PARENT logins for all schools/students.
Writes login-credentials-all.xlsx to the repo root.

Login format:
  Principal : {siteCode}@edubeam.com        / {siteCode}
  Student   : st{siteCode}{admNo}@edubeam.com / st{siteCode}{admNo}
  Parent    : pr{siteCode}{admNo}@edubeam.com / pr{siteCode}{admNo}

Rationale: admissionNo is unique only within a school, not globally.
Including the siteCode guarantees uniqueness across all 500 schools.

Usage:
    python generate-logins.py
"""

import sqlite3
import sys
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

# Force UTF-8 output on Windows
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def _pip(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", *pkgs],
                          stdout=subprocess.DEVNULL)

try:
    import bcrypt as _bcrypt
except ImportError:
    print("Installing bcrypt..."); _pip("bcrypt"); import bcrypt as _bcrypt

try:
    import openpyxl as _ox
except ImportError:
    print("Installing openpyxl..."); _pip("openpyxl"); import openpyxl as _ox

# Config
SCRIPT_DIR   = Path(__file__).resolve().parent
DB_PATH      = SCRIPT_DIR / "packages" / "db" / "prisma" / "dev.db"
OUT_PATH     = SCRIPT_DIR / "login-credentials-all.xlsx"
DOMAIN       = "edubeam.com"
TENANT_ID    = "t_uk"
BCRYPT_COST  = 6
HASH_WORKERS = 16
BATCH_SIZE   = 500

NOW = datetime.now(timezone.utc).isoformat()


def _hash(password: str) -> str:
    return _bcrypt.hashpw(password.encode(), _bcrypt.gensalt(rounds=BCRYPT_COST)).decode()


def hash_batch(passwords: list) -> list:
    with ThreadPoolExecutor(max_workers=HASH_WORKERS) as ex:
        return list(ex.map(_hash, passwords))


INSERT_SQL = """
INSERT OR IGNORE INTO "User"
  (id, email, "passwordHash", name, role, "tenantId",
   "districtId", "blockId", "schoolId", "studentId", "linkedStudentIds",
   active, "createdAt")
VALUES (?,?,?,?,?,?,?,NULL,?,?,?,1,?)
"""


def main():
    print()
    print("===========================================================")
    print(" Edubeam LMS -- Bulk Login Generator")
    print("===========================================================")
    print(f" DB      : {DB_PATH}")
    print(f" Output  : {OUT_PATH}")
    print(f" Workers : {HASH_WORKERS}  |  bcrypt cost : {BCRYPT_COST}")
    print("-----------------------------------------------------------")
    print()

    if not DB_PATH.exists():
        print(f"ERROR: database not found at {DB_PATH}"); sys.exit(1)

    conn = sqlite3.connect(str(DB_PATH), timeout=60)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=60000")
    cur = conn.cursor()

    # ------------------------------------------------------------------
    # 0. Clean up any previously generated STUDENT / PARENT users so we
    #    can regenerate them with the correct siteCode-prefixed format.
    # ------------------------------------------------------------------
    print("> Cleaning up old STUDENT/PARENT user rows...")
    cur.execute('DELETE FROM "User" WHERE role IN (\'STUDENT\', \'PARENT\')')
    deleted = cur.execute("SELECT changes()").fetchone()[0]
    conn.commit()
    print(f"  Removed {deleted:,} old rows.")
    print()

    # ------------------------------------------------------------------
    # 1. Load schools with their siteCode
    # ------------------------------------------------------------------
    print("> Loading schools...")
    cur.execute("""
        SELECT s.id, s.name, s."siteCode", s."udiseCode", b."districtId"
        FROM   School s
        LEFT JOIN Block b ON s."blockId" = b.id
        ORDER  BY s.name
    """)
    schools_raw = cur.fetchall()
    print(f"  {len(schools_raw)} schools found")

    # Build siteCode lookup: schoolId -> siteCode
    school_site = {}
    for sid, name, site_code, udise, dist_id in schools_raw:
        school_site[sid] = (site_code or f"s{udise}").lower().strip()

    school_excel = []
    pw_list  = []
    meta_list = []

    for sid, name, site_code, udise, district_id in schools_raw:
        local = school_site[sid]
        email = f"{local}@{DOMAIN}"
        pw_list.append(local)
        meta_list.append((f"us_{local}", email, f"Principal -- {name}",
                          "PRINCIPAL", district_id, sid, name, district_id or ""))

    print(f"  Hashing {len(pw_list)} school passwords...")
    hashes = hash_batch(pw_list)

    rows = []
    for (uid, email, disp, role, district_id, school_id, sname, dist_name), pw, ph in \
            zip(meta_list, pw_list, hashes):
        rows.append((uid, email, ph, disp, role, TENANT_ID,
                     district_id, school_id, None, None, NOW))
        school_excel.append(["Principal", sname, dist_name, email, pw])

    cur.executemany(INSERT_SQL, rows)
    conn.commit()
    print(f"  OK: {len(rows)} principal logins created/updated")
    print()

    # ------------------------------------------------------------------
    # 2. Students + Parents  (key = siteCode + admissionNo)
    # ------------------------------------------------------------------
    print("> Loading students...")
    cur.execute("""
        SELECT st.id, st.name, st."admissionNo", st."rollNo",
               st.grade, st.section, st."schoolId", st."guardianName",
               s.name AS school_name, b."districtId"
        FROM   Student st
        LEFT JOIN School s ON st."schoolId" = s.id
        LEFT JOIN Block  b ON s."blockId"   = b.id
        ORDER  BY s.name, st.grade
    """)
    students = cur.fetchall()
    total = len(students)
    print(f"  {total:,} students found")

    student_excel = []
    parent_excel  = []
    done = 0
    # Track (siteCode, rawAdmKey) -> count, to deduplicate within a school
    seen_keys: dict = {}

    for i in range(0, total, BATCH_SIZE):
        chunk = students[i : i + BATCH_SIZE]

        st_pws, pr_pws = [], []
        for (sid, sname, adm, roll, grade, section,
             school_id, guardian, school_name, district_id) in chunk:
            site = school_site.get(school_id, "school")
            # Sanitize the admission number
            raw_key = (adm or roll or "").replace(" ", "").lower()
            BAD = {"#","##","###","####","#####","na","n/a","nil","none","0",""}
            if raw_key in BAD:
                raw_key = sid[:8]
            # Deduplicate within school: append b/c/d/... for 2nd/3rd/... occurrence
            dedup_key = (site, raw_key)
            count = seen_keys.get(dedup_key, 0)
            seen_keys[dedup_key] = count + 1
            suffix = "" if count == 0 else chr(ord('b') + count - 1)
            key = f"{site}{raw_key}{suffix}"
            st_pws.append(f"st{key}")
            pr_pws.append(f"pr{key}")

        st_hashes = hash_batch(st_pws)
        pr_hashes = hash_batch(pr_pws)

        st_rows, pr_rows = [], []
        for j, (sid, sname, adm, roll, grade, section,
                school_id, guardian, school_name, district_id) in enumerate(chunk):
            grade_label = f"Class {grade}{'-'+section if section else ''}"

            st_local = st_pws[j]
            st_uid   = f"ustu_{st_local[2:]}"   # strip "st" prefix for id
            st_email = f"{st_local}@{DOMAIN}"
            st_rows.append((st_uid, st_email, st_hashes[j], sname, "STUDENT", TENANT_ID,
                            district_id, school_id, sid, None, NOW))
            student_excel.append(["Student", sname, school_name or "",
                                   grade_label, st_email, st_local])

            pr_local = pr_pws[j]
            pr_uid   = f"upar_{pr_local[2:]}"   # strip "pr" prefix for id
            pr_email = f"{pr_local}@{DOMAIN}"
            pr_rows.append((pr_uid, pr_email, pr_hashes[j], f"Parent -- {sname}", "PARENT",
                            TENANT_ID, district_id, school_id, None, sid, NOW))
            parent_excel.append(["Parent", sname, school_name or "",
                                  guardian or "", pr_email, pr_local])

        cur.executemany(INSERT_SQL, st_rows)
        cur.executemany(INSERT_SQL, pr_rows)
        conn.commit()

        done += len(chunk)
        pct = done * 100 // total
        print(f"  Students+Parents: {done:,} / {total:,}  ({pct}%)")

    conn.close()
    print(f"  OK: {len(student_excel):,} student logins created/updated")
    print(f"  OK: {len(parent_excel):,} parent logins created/updated")
    print()

    # ------------------------------------------------------------------
    # 3. Excel workbook
    # ------------------------------------------------------------------
    print("> Generating Excel workbook...")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    def header_style(ws, cols, bg_hex):
        hf   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        hfil = PatternFill("solid", fgColor=bg_hex)
        for c in range(1, cols + 1):
            cell = ws.cell(1, c)
            cell.font      = hf
            cell.fill      = hfil
            cell.alignment = Alignment(horizontal="center")

    wb = Workbook()

    # Summary
    ws0 = wb.active; ws0.title = "Summary"
    ws0.append(["Edubeam LMS -- All Login Credentials"])
    ws0.append([f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"])
    ws0.append([])
    ws0.append(["Category", "Count", "Username / Email format", "Password rule"])
    ws0.append(["School / Principal", len(school_excel),
                "{siteCode}@edubeam.com", "Same as username (part before @)"])
    ws0.append(["Student", len(student_excel),
                "st{siteCode}{admNo}@edubeam.com", "Same as username"])
    ws0.append(["Parent", len(parent_excel),
                "pr{siteCode}{admNo}@edubeam.com", "Same as username"])
    ws0.append([])
    ws0.append(["TOTAL", len(school_excel)+len(student_excel)+len(parent_excel)])
    header_style(ws0, 4, "1F4E79")
    ws0.column_dimensions["A"].width = 26
    ws0.column_dimensions["B"].width = 12
    ws0.column_dimensions["C"].width = 40
    ws0.column_dimensions["D"].width = 32

    # School Principals
    ws1 = wb.create_sheet("School Principals")
    ws1.append(["Role", "School Name", "District", "Email (Username)", "Password"])
    for r in school_excel: ws1.append(r)
    header_style(ws1, 5, "2E75B6")
    for w, col in zip([12, 45, 22, 38, 22], ["A", "B", "C", "D", "E"]):
        ws1.column_dimensions[col].width = w

    # Students
    ws2 = wb.create_sheet("Students")
    ws2.append(["Role", "Student Name", "School", "Grade", "Email (Username)", "Password"])
    for r in student_excel: ws2.append(r)
    header_style(ws2, 6, "375623")
    for w, col in zip([10, 32, 42, 14, 48, 32], ["A", "B", "C", "D", "E", "F"]):
        ws2.column_dimensions[col].width = w

    # Parents
    ws3 = wb.create_sheet("Parents")
    ws3.append(["Role", "Parent of", "School", "Guardian Name", "Email (Username)", "Password"])
    for r in parent_excel: ws3.append(r)
    header_style(ws3, 6, "7B3F00")
    for w, col in zip([10, 32, 42, 28, 48, 32], ["A", "B", "C", "D", "E", "F"]):
        ws3.column_dimensions[col].width = w

    wb.save(str(OUT_PATH))

    n_total = len(school_excel) + len(student_excel) + len(parent_excel)
    print()
    print("===========================================================")
    print(" DONE!")
    print(f"    Principal logins : {len(school_excel):,}")
    print(f"    Student logins   : {len(student_excel):,}")
    print(f"    Parent logins    : {len(parent_excel):,}")
    print(f"    Total            : {n_total:,}")
    print(f"    Excel file       : {OUT_PATH}")
    print("===========================================================")
    print()


if __name__ == "__main__":
    main()
